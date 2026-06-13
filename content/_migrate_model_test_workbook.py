#!/usr/bin/env python3
"""
_migrate_model_test_workbook.py — ONE-TIME migration.

Freezes the validated model-test build (editor's raw 5 sheets joined to the bank,
+ my translations of the new questions) into a clean, stable, ID-keyed source:

    content/sources/model_test_workbook.xlsx

After this, build_model_tests.py reads THIS workbook (deterministic, keyed by
question_id) instead of the gitignored editor file + fuzzy join. To edit a test,
change a cell here.

Reads the already-built src/data/json/model_tests.json and
model_test_questions.json. Run once:

    python3 content/_migrate_model_test_workbook.py
"""

import json
import os

import openpyxl
from openpyxl.styles import Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
TESTS = os.path.join(ROOT, "src", "data", "json", "model_tests.json")
NEWQ = os.path.join(ROOT, "src", "data", "json", "model_test_questions.json")
OUT = os.path.join(HERE, "sources", "model_test_workbook.xlsx")

HEAD_FILL = PatternFill("solid", fgColor="FFE600")
HEAD_FONT = Font(bold=True)


def header(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
    ws.freeze_panes = "A2"


def main():
    tests = json.load(open(TESTS, encoding="utf-8"))
    newq = json.load(open(NEWQ, encoding="utf-8"))

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Tests"
    header(ws, ["test_id", "title_fi", "title_en", "time_minutes", "pass_mark", "order"])
    for i, t in enumerate(tests, start=1):
        ws.append([t["id"], t["title_fi"], t["title_en"], t["time_minutes"], t["pass_mark"], i])

    ws = wb.create_sheet("Test_Questions")
    header(ws, ["test_id", "slot", "question_id"])
    for t in tests:
        for slot, qid in enumerate(t["question_ids"], start=1):
            ws.append([t["id"], slot, qid])

    ws = wb.create_sheet("New_Questions")
    header(ws, ["id", "correct_option", "question_fi", "question_en",
                "option_a_fi", "option_a_en", "option_b_fi", "option_b_en",
                "option_c_fi", "option_c_en", "explanation_en"])
    by_key = {o["key"]: o for q in newq for o in []}  # noop placeholder for clarity
    for q in newq:
        opt = {o["key"]: o for o in q["options"]}
        ws.append([
            q["id"], q["correct_option"], q["question"]["fi"], q["question"]["en"],
            opt["A"]["fi"], opt["A"]["en"], opt["B"]["fi"], opt["B"]["en"],
            opt["C"]["fi"], opt["C"]["en"], q["explanation_en"],
        ])

    for sheet, widths in {
        "Tests": {"A": 10, "B": 18, "C": 16, "D": 13, "E": 11, "F": 7},
        "Test_Questions": {"A": 10, "B": 7, "C": 12},
        "New_Questions": {"A": 10, "B": 14, "C": 70, "D": 70, "E": 45, "F": 45,
                          "G": 45, "H": 45, "I": 45, "J": 45, "K": 70},
    }.items():
        for col, w in widths.items():
            wb[sheet].column_dimensions[col].width = w

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  Tests        : {len(tests)}")
    print(f"  Test slots   : {sum(len(t['question_ids']) for t in tests)}")
    print(f"  New questions: {len(newq)}")


if __name__ == "__main__":
    main()
