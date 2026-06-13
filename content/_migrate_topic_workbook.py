#!/usr/bin/env python3
"""
_migrate_topic_workbook.py — ONE-TIME migration.

Freezes the validated fuzzy-join (editor's messy Excel → shipped questions.json)
into a clean, stable, ID-keyed source workbook:

    content/sources/topic_workbook.xlsx

After this runs, build_topics.py reads THIS workbook (deterministic, keyed by
question_id) instead of re-doing the fragile text match against the editor file.
The editor's original file is kept only as the provenance of this migration.

Reads the already-built src/data/json/topic_practice.json (the validated join)
plus questions.json for human-readable Finnish text. Run once:

    python3 content/_migrate_topic_workbook.py
"""

import json
import os

import openpyxl
from openpyxl.styles import Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
TP = os.path.join(ROOT, "src", "data", "json", "topic_practice.json")
QUESTIONS = os.path.join(ROOT, "src", "data", "json", "questions.json")
OUT = os.path.join(HERE, "sources", "topic_workbook.xlsx")

HEAD_FILL = PatternFill("solid", fgColor="FFE600")
HEAD_FONT = Font(bold=True)


def header(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font = HEAD_FONT
        c.fill = HEAD_FILL


def main():
    tp = json.load(open(TP, encoding="utf-8"))
    questions = {q["id"]: q for q in json.load(open(QUESTIONS, encoding="utf-8"))}

    wb = openpyxl.Workbook()

    # ── Sections ─────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Sections"
    header(ws, ["section_id", "name_fi", "name_en", "description", "pass_correct", "pass_total", "order"])
    for s in tp["sections"]:
        ws.append([s["id"], s["name_fi"], s["name_en"], s["description"],
                   s["pass_correct"], s["pass_total"], s["order"]])

    # ── Lessons ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Lessons")
    header(ws, ["lesson_id", "section_id", "lesson_name", "order"])
    for l in tp["lessons"]:
        ws.append([l["id"], l["section_id"], l["name"], l["order"]])

    # ── Questions (the actual mapping — keyed by stable question_id) ──────────
    ws = wb.create_sheet("Questions")
    header(ws, ["question_id", "section_id", "lesson_id", "lesson_name", "order_in_lesson", "question_fi"])
    for l in tp["lessons"]:
        for i, qid in enumerate(l["question_ids"], start=1):
            fi = (questions.get(qid, {}).get("question", {}) or {}).get("fi", "")
            ws.append([qid, l["section_id"], l["id"], l["name"], i, fi])

    # column widths for editor readability
    for sheet, widths in {
        "Sections": {"A": 18, "B": 52, "C": 42, "D": 80, "E": 13, "F": 11, "G": 7},
        "Lessons": {"A": 44, "B": 18, "C": 40, "D": 7},
        "Questions": {"A": 12, "B": 18, "C": 44, "D": 40, "E": 14, "F": 90},
    }.items():
        for col, w in widths.items():
            wb[sheet].column_dimensions[col].width = w
        wb[sheet].freeze_panes = "A2"

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  Sections : {len(tp['sections'])}")
    print(f"  Lessons  : {len(tp['lessons'])}")
    print(f"  Questions: {sum(len(l['question_ids']) for l in tp['lessons'])}")


if __name__ == "__main__":
    main()
