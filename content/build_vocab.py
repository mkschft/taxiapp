#!/usr/bin/env python3
"""
build_vocab.py — compile the vocabulary workbook into the app's vocab data.

Mirrors the new Vocabulary flow:
    Vocabulary  →  practice sets  →  each Set has a Lesson (word cards) + a Quiz

Input:
  content/sources/vocab_workbook.xlsx   (SOURCE OF TRUTH, from the content editor)
    Sheet "Vocabulary_Groups_Rearranged"  → lesson word cards
    Sheet "Quiz set"                       → quiz questions

Output:
  src/data/json/vocab.json     { sets, words, questions } — the app reads this
  content/output/_vocab_report.txt   build + validation report

Cleanup applied here (the Excel was AI-generated, so normalization lives in the
pipeline — the editor's file is never hand-edited):
  * Sets 10-11 (Conjunctions, WH-words) are EXCLUDED. Their lesson rows are
    English category titles, not Finnish word cards — they belong to the
    Clue Word section. Their raw rows remain in the Excel, untouched.
  * `forms` cells are parsed from three mixed formats: "fi = en",
    "fi (gloss) fi2 (gloss2)", and bare inflections ("tilanteen").
  * Duplicate forms removed; a few obvious typos fixed (see TYPO_FIX).
  * Quiz sheet is joined to lesson sets BY POSITION (groups appear in set
    order 1..N), because the quiz sheet labels sets by name only and a couple
    of names differ slightly from the lesson sheet.

Run from the project root:  python3 content/build_vocab.py
"""

import json
import os
import re

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
WORKBOOK = os.path.join(HERE, "sources", "vocab_workbook.xlsx")
OUT_JSON = os.path.join(ROOT, "src", "data", "json", "vocab.json")
OUT_REPORT = os.path.join(HERE, "output", "_vocab_report.txt")

LESSON_SHEET = "Vocabulary_Groups_Rearranged"
QUIZ_SHEET = "Quiz set"

# Sets to exclude from vocabulary. (Sets 10 & 11 — Conjunctions and WH-words —
# are now kept, matching the Excel's 11 sets.)
EXCLUDE_SETS = set()

# Sets 10/11 list a different sub-page name on every row; give them a clean name.
SET_NAME_OVERRIDE = {
    10: "Conjunctions",
    11: "WH / Question Words",
}

OPTION_KEYS = ("A", "B", "C")

# Best-effort theme link: vocab set_no → exam category id (categories.json).
CATEGORY_BY_SET = {
    1: "special_needs",      # Passenger Assistance & Accessibility
    2: "special_needs",      # Passenger Assistance & Communication
    3: "passenger_safety",   # Child, Seatbelt & School Transport Safety
    4: "traffic_safety",     # Traffic Safety & Driving Rules
    5: "traffic_safety",     # Police, Emergency & Risky Behaviour
    6: "customer_service",   # Payment, Receipt, Kela & Fare
    7: "customer_service",   # Trip, Route, Luggage & Useful Action Words
    8: "traffic_safety",     # Licence, Duties & Legal Rules
    9: "customer_service",   # Passenger Service, Communication & Complaints
}

# Small, explicit corrections for known AI-generated typos in the FI source.
TYPO_FIX = {
    "poispain": "poispäin",
}


def clean(s):
    if s is None:
        return None
    s = str(s).strip()
    return s or None


def parse_page_count(pc):
    """'3/8' -> (3, 8). Falls back to (None, None)."""
    m = re.match(r"\s*(\d+)\s*/\s*(\d+)", str(pc or ""))
    if not m:
        return None, None
    return int(m.group(1)), int(m.group(2))


def _split_top(s, sep):
    """Split on `sep`, but only at paren depth 0 (sets 10/11 glosses contain
    '(... ; ex: ...)' with semicolons and '=' inside the parentheses)."""
    parts, depth, cur = [], 0, ""
    for ch in s:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth = max(0, depth - 1)
        if ch == sep and depth == 0:
            parts.append(cur)
            cur = ""
        else:
            cur += ch
    if cur.strip():
        parts.append(cur)
    return parts


def _find_top_eq(s):
    """Index of the first '=' at paren depth 0, or -1."""
    depth = 0
    for i, ch in enumerate(s):
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth = max(0, depth - 1)
        elif ch == "=" and depth == 0:
            return i
    return -1


def parse_forms(raw):
    """Parse the forms_suffix_variations cell into [{fi, en}].

    Handles the formats found across all 11 sets, possibly mixed in one cell:
      A) 'kieltäytyä = to refuse; kieltää = to forbid'
      B) 'puhelin (phone) puhelimeen (into/on phone)'
      C) bare inflection with no gloss: 'tilanteen'
      D) sets 10/11: 'ja = and (both connected ideas matter; ex: ... = ...); sekä = ...'
         — the meaning is the text before '(' ; the parenthetical is extra context.
    """
    if not raw:
        return []
    out = []
    for part in _split_top(str(raw), ";"):
        part = part.strip()
        if not part:
            continue
        eq = _find_top_eq(part)
        if eq >= 0:
            fi = part[:eq].strip()
            en = part[eq + 1:].strip()
            paren = en.find("(")            # drop trailing "(...context...)"
            if paren >= 0:
                en = en[:paren].strip()
            out.append({"fi": fi, "en": en})
        elif "(" in part:
            for m in re.finditer(r"([^()]+?)\s*\(([^)]*)\)", part):
                fi = re.sub(r"\s+", " ", m.group(1)).strip(" ,")
                en = m.group(2).strip()
                if fi:
                    out.append({"fi": fi, "en": en})
        else:
            out.append({"fi": part, "en": ""})               # bare inflection

    # typo fix + dedupe (case-insensitive on fi+en)
    seen, deduped = set(), []
    for f in out:
        f["fi"] = TYPO_FIX.get(f["fi"].lower(), f["fi"])
        if not f["fi"]:
            continue
        key = (f["fi"].lower(), f["en"].lower())
        if key in seen:
            continue
        seen.add(key)
        deduped.append(f)
    return deduped


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    lesson_rows = [r for r in wb[LESSON_SHEET].iter_rows(min_row=2, values_only=True)
                   if r[0] is not None]
    quiz_rows = [r for r in wb[QUIZ_SHEET].iter_rows(min_row=2, values_only=True)
                 if (r[0] is not None or r[1] is not None)]

    report = []

    # ---- Lesson words, grouped by set_no -----------------------------------
    sets = {}        # set_no -> set meta (name, order)
    words = []       # VocabLessonWord[]
    for r in lesson_rows:
        set_no, name, page_count, base_word, meaning, forms, exam_use = r[:7]
        set_no = int(set_no)
        if set_no in EXCLUDE_SETS:
            continue
        if set_no not in sets:
            sets[set_no] = {"name": SET_NAME_OVERRIDE.get(set_no, clean(name))}
        index, total = parse_page_count(page_count)
        words.append({
            "id": f"set-{set_no}-w{index}",
            "set_id": f"set-{set_no}",
            "index": index,
            "total_in_set": total,
            "word_fi": clean(base_word),
            "meaning_en": clean(meaning),
            "forms_fi": parse_forms(forms),
            "exam_use_en": clean(exam_use),
        })

    # ---- Quiz questions: join to sets BY POSITION --------------------------
    # Distinct quiz group names in first-seen order == set order 1..N.
    quiz_group_order = list(dict.fromkeys(r[0] for r in quiz_rows))
    included_set_nos = sorted(sets)  # 1..9 in order
    if len(quiz_group_order) < len(included_set_nos):
        report.append(f"[WARN] quiz groups ({len(quiz_group_order)}) < lesson sets "
                      f"({len(included_set_nos)}) — position join may be off")
    # Map each quiz group name -> set_no by position (skip excluded sets' groups).
    name_to_set_no = {}
    for pos, gname in enumerate(quiz_group_order, start=1):
        if pos in EXCLUDE_SETS:
            continue
        name_to_set_no[gname] = pos

    # Index lesson words per set for the optional soft link.
    words_by_set = {}
    for w in words:
        words_by_set.setdefault(w["set_id"], {})[(w["word_fi"] or "").lower()] = w["id"]

    questions = []
    for r in quiz_rows:
        gname, order, q_word, correct_meaning, a, b, c, correct_opt = r[:8]
        set_no = name_to_set_no.get(gname)
        if set_no is None:           # belongs to an excluded set
            continue
        set_id = f"set-{set_no}"
        opts = {"A": clean(a), "B": clean(b), "C": clean(c)}
        co = (clean(correct_opt) or "").upper()
        if co not in OPTION_KEYS:
            report.append(f"[WARN] {set_id} q{order}: bad correct_option {correct_opt!r}")
        lesson_word_id = words_by_set.get(set_id, {}).get((clean(q_word) or "").lower())
        questions.append({
            "id": f"{set_id}-q{int(order)}",
            "set_id": set_id,
            "index": int(order),
            "prompt_word_fi": clean(q_word),
            "options": [{"key": k, "en": opts[k]} for k in OPTION_KEYS],
            "correct_option": co,
            "correct_meaning_en": clean(correct_meaning),
            "lesson_word_id": lesson_word_id,   # may be None
        })

    # ---- Assemble VocabSet rows (with derived counts) ----------------------
    vocab_sets = []
    for order, set_no in enumerate(included_set_nos, start=1):
        wc = sum(1 for w in words if w["set_id"] == f"set-{set_no}")
        qc = sum(1 for q in questions if q["set_id"] == f"set-{set_no}")
        vocab_sets.append({
            "id": f"set-{set_no}",
            "set_no": set_no,
            "name": sets[set_no]["name"],
            "category_id": CATEGORY_BY_SET.get(set_no),
            "order": order,
            "word_count": wc,
            "question_count": qc,
        })

    out = {"sets": vocab_sets, "words": words, "questions": questions}
    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    # ---- Report ------------------------------------------------------------
    report.insert(0, "VOCAB BUILD REPORT")
    report.insert(1, "=" * 50)
    report.append(f"sets:      {len(vocab_sets)}")
    report.append(f"words:     {len(words)}")
    report.append(f"questions: {len(questions)}")
    report.append(f"excluded sets (clue-word, sent to Clue Word section): "
                  f"{sorted(EXCLUDE_SETS)}")
    report.append("")
    report.append("per-set:")
    for s in vocab_sets:
        report.append(f"  {s['id']:>7}  {s['word_count']:>2}w / {s['question_count']:>2}q  "
                      f"[{s['category_id'] or '-'}]  {s['name']}")
    linked = sum(1 for q in questions if q["lesson_word_id"])
    report.append("")
    report.append(f"quiz→lesson soft links resolved: {linked}/{len(questions)}")

    os.makedirs(os.path.dirname(OUT_REPORT), exist_ok=True)
    with open(OUT_REPORT, "w", encoding="utf-8") as f:
        f.write("\n".join(report) + "\n")

    print("\n".join(report))
    print(f"\nwrote {OUT_JSON}")
    print(f"wrote {OUT_REPORT}")


if __name__ == "__main__":
    main()
