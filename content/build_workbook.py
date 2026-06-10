#!/usr/bin/env python3
"""
build_workbook.py — BOOTSTRAP / RESET ONLY.

Generates content/content_workbook.xlsx (the content editor's working file)
from the current src/data/json/questions.json.

⚠️  This OVERWRITES content_workbook.xlsx. Only run it to create the workbook
    for the first time, or to reset it from the compiled app data. Day-to-day
    you EDIT the workbook by hand and run build_appready.py — you do NOT run this.

Usage:  python3 content/build_workbook.py
"""

import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
QUESTIONS = os.path.join(ROOT, "src", "data", "json", "questions.json")
WORKBOOK = os.path.join(HERE, "content_workbook.xlsx")
OPTION_KEYS = ["A", "B", "C"]

# (column key, header label, group) — group drives the header colour
COLS = [
    ("question_id",   "Question ID",   "ref"),
    ("section",       "Category",      "ref"),
    ("lesson",        "Topic",         "ref"),
    ("ref_no",        "Ref #",         "ref"),
    ("question_fi",   "Question (FI)", "fi"),
    ("question_en",   "Question (EN)", "en"),
    ("option_a_fi",   "Option A (FI)", "fi"),
    ("option_a_en",   "Option A (EN)", "en"),
    ("option_b_fi",   "Option B (FI)", "fi"),
    ("option_b_en",   "Option B (EN)", "en"),
    ("option_c_fi",   "Option C (FI)", "fi"),
    ("option_c_en",   "Option C (EN)", "en"),
    ("correct_option","Correct (A/B/C)", "key"),
    ("fw_words",      "Focus words (FI)",     "fw"),
    ("fw_meanings",   "Focus words (EN)",     "fw"),
    ("pcw_words",     "Positive clues (FI)",  "pcw"),
    ("pcw_meanings",  "Positive clues (EN)",  "pcw"),
    ("ncw_words",     "Negative clues (FI)",  "ncw"),
    ("ncw_meanings",  "Negative clues (EN)",  "ncw"),
    ("explanation_en","Explanation (EN)", "meta"),
    ("difficulty",    "Difficulty",       "meta"),
    ("tags",          "Tags",             "meta"),
    ("status",        "Status",           "meta"),
    ("reviewer_notes","Reviewer notes",   "meta"),
]

GROUP_FILL = {
    "ref":  "D9D9D9",  # grey  — reference, do not edit
    "fi":   "DDEBF7",  # blue  — Finnish text
    "en":   "E2EFDA",  # green — English text
    "key":  "FFF2CC",  # gold  — correct answer
    "fw":   "FFF2CC",  # gold  — focus words
    "pcw":  "E2EFDA",  # green — positive clues
    "ncw":  "FCE4D6",  # red   — negative clues
    "meta": "EDEDED",  # light grey — meta
}

WIDTHS = {
    "question_id": 12, "section": 20, "lesson": 22, "ref_no": 8,
    "correct_option": 13, "difficulty": 12, "status": 14,
}
DEFAULT_WIDTH = 34


def join_clues(annotations, ctype, field):
    items = [a for a in annotations if a.get("clue_type") == ctype]
    if field == "fi":
        return "; ".join(a["text_fi"] for a in items if a.get("text_fi"))
    return "; ".join((a.get("meaning_en") or "") for a in items)


def row_values(q):
    opts = {o["key"]: o for o in q["options"]}
    ann = q.get("clue_annotations", [])
    return {
        "question_id": q["id"],
        "section": q.get("category_en"),
        "lesson": q.get("source_topic_fi"),
        "ref_no": q.get("ref_no"),
        "question_fi": q["question"].get("fi"),
        "question_en": q["question"].get("en"),
        "option_a_fi": opts.get("A", {}).get("fi"),
        "option_a_en": opts.get("A", {}).get("en"),
        "option_b_fi": opts.get("B", {}).get("fi"),
        "option_b_en": opts.get("B", {}).get("en"),
        "option_c_fi": opts.get("C", {}).get("fi"),
        "option_c_en": opts.get("C", {}).get("en"),
        "correct_option": q.get("correct_option"),
        "fw_words": join_clues(ann, "fw", "fi"),
        "fw_meanings": join_clues(ann, "fw", "en"),
        "pcw_words": join_clues(ann, "pcw", "fi"),
        "pcw_meanings": join_clues(ann, "pcw", "en"),
        "ncw_words": join_clues(ann, "ncw", "fi"),
        "ncw_meanings": join_clues(ann, "ncw", "en"),
        "explanation_en": q.get("explanation_en"),
        "difficulty": q.get("difficulty"),
        "tags": "; ".join(q.get("tags", [])),
        "status": q.get("status"),
        "reviewer_notes": q.get("reviewer_notes"),
    }


def build_readme(wb):
    ws = wb.create_sheet("How to use", 0)
    ws.sheet_view.showGridLines = False
    lines = [
        ("HOW TO EDIT THIS WORKBOOK", True),
        ("", False),
        ("This is the single source of all question content for the app.", False),
        ("Edit the 'Questions' sheet, save, and hand the file to the developer", False),
        ("(or run the build yourself — see the guide). Each saved + built change", False),
        ("updates the app.", False),
        ("", False),
        ("COLUMN COLOURS", True),
        ("  Grey   = reference only — DO NOT edit (Question ID, Category, Topic, Ref #)", False),
        ("  Blue   = Finnish text shown in the app", False),
        ("  Green  = English translation shown when the user taps 'Translate'", False),
        ("  Gold   = the correct answer + focus words (yellow highlight)", False),
        ("  Red    = negative clue words (point AWAY from the answer)", False),
        ("", False),
        ("THE 3 CLUE TYPES", True),
        ("  Focus words  — key Finnish words to highlight yellow in the question", False),
        ("  Positive clues — words in the CORRECT option that signal it is right", False),
        ("  Negative clues — words in a WRONG option that signal it is wrong", False),
        ("", False),
        ("CLUE WORD RULES (important)", True),
        ("  • Separate multiple words with a semicolon ;   e.g.  avustaa; auttaa", False),
        ("  • Each FI clue word MUST appear EXACTLY as written in the FI text,", False),
        ("    otherwise the app cannot highlight it (the build will warn you).", False),
        ("  • The EN meanings line up by position with the FI words", False),
        ("    (1st meaning = 1st word, 2nd = 2nd, ...).", False),
        ("", False),
        ("GOLDEN RULES", True),
        ("  • Never change a Question ID.", False),
        ("  • 'Correct (A/B/C)' must be a single letter: A, B or C.", False),
        ("  • 'Difficulty' must be Easy, Medium or Hard.", False),
        ("  • Leave a row's English blank only if it still needs writing.", False),
        ("  • Use 'Reviewer notes' to flag anything you're unsure about.", False),
        ("", False),
        ("See content/CONTENT_GUIDE.md for the full step-by-step guide.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=13 if bold else 11,
                      color="1F4E78" if bold else "000000")
    ws.column_dimensions["A"].width = 90


def main():
    with open(QUESTIONS, encoding="utf-8") as f:
        questions = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # header
    for ci, (key, label, group) in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = Font(bold=True, color="000000", size=11)
        c.fill = PatternFill("solid", fgColor=GROUP_FILL[group])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(key, DEFAULT_WIDTH)

    # rows
    for ri, q in enumerate(questions, start=2):
        vals = row_values(q)
        for ci, (key, _label, group) in enumerate(COLS, start=1):
            c = ws.cell(row=ri, column=ci, value=vals.get(key))
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
            if group == "ref":
                c.fill = PatternFill("solid", fgColor="F2F2F2")

    # freeze header + first column
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30

    n = len(questions) + 1
    # dropdowns
    diff_col = get_column_letter([k for k, *_ in COLS].index("difficulty") + 1)
    corr_col = get_column_letter([k for k, *_ in COLS].index("correct_option") + 1)
    dv_diff = DataValidation(type="list", formula1='"Easy,Medium,Hard"', allow_blank=True)
    dv_corr = DataValidation(type="list", formula1='"A,B,C"', allow_blank=True)
    ws.add_data_validation(dv_diff)
    ws.add_data_validation(dv_corr)
    dv_diff.add(f"{diff_col}2:{diff_col}{n}")
    dv_corr.add(f"{corr_col}2:{corr_col}{n}")

    build_readme(wb)
    wb.save(WORKBOOK)
    print(f"Wrote {WORKBOOK}")
    print(f"  {len(questions)} question rows + 'How to use' sheet")
    print("  ⚠️  This is a reset. Day-to-day, edit the workbook and run build_appready.py")


if __name__ == "__main__":
    main()
