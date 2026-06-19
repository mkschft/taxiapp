#!/usr/bin/env python3
"""
build_topics.py — compile the Topic Practice section.

Flow:
    Topic Practice  →  4 Section cards
                    →  each Section has Lessons (sub-categories)
                    →  each Lesson has its own questions + quiz

Input (SOURCE OF TRUTH):
  content/sources/topic_workbook.xlsx   clean, ID-keyed taxonomy
    'Sections'   section_id | name_fi | name_en | description | pass_correct | pass_total | order
    'Lessons'    lesson_id | section_id | lesson_name | order
    'Questions'  question_id | section_id | lesson_id | lesson_name | order_in_lesson | question_fi

  src/data/json/questions.json          the shipped question bank — question
                                        CONTENT lives here; this workbook only
                                        maps each question_id to a section+lesson.

This workbook was frozen from a validated fuzzy-join of the content editor's
original Excel (see content/_migrate_topic_workbook.py). It is now keyed by the
stable question_id, so this build is fully deterministic — to re-tag a question,
edit the 'Questions' sheet, not text. question_fi is there for editor readability
only; it is never used to match.

Output:
  src/data/json/topic_practice.json     { sections, lessons } — lessons carry
                                         ordered question_ids[] into questions.json
  content/output/_topics_report.txt      build summary + any integrity warnings

Run from the project root:  python3 content/build_topics.py
"""

import json
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
WORKBOOK = os.path.join(HERE, "sources", "topic_workbook.xlsx")
QUESTIONS = os.path.join(ROOT, "src", "data", "json", "questions.json")
OUT_JSON = os.path.join(ROOT, "src", "data", "json", "topic_practice.json")
OUT_REPORT = os.path.join(HERE, "output", "_topics_report.txt")


def rows(ws):
    head = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        yield dict(zip(head, r))


def _usable(q):
    """A question is usable in practice only if it has exactly one correct
    option and isn't a flagged 'source-unclear' placeholder. Filtering here
    keeps empty/ungradeable questions (e.g. Q164/Q166/Q171/Q178) out of the
    study flow — the same guard the model-test builder applies."""
    if q.get("status") == "source-unclear":
        return False
    return sum(1 for o in q.get("options", []) if o.get("is_correct")) == 1


def main():
    bank = json.load(open(QUESTIONS, encoding="utf-8"))
    valid_ids = {q["id"] for q in bank if _usable(q)}
    unusable_ids = {q["id"] for q in bank if not _usable(q)}
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    warnings = []

    # ── Sections ─────────────────────────────────────────────────────────────
    sections = []
    for r in rows(wb["Sections"]):
        sections.append({
            "id": r["section_id"], "category_id": r["section_id"],
            "name_fi": r["name_fi"], "name_en": r["name_en"], "description": r["description"],
            "pass_correct": r["pass_correct"], "pass_total": r["pass_total"],
            "order": r["order"], "lesson_count": 0, "question_count": 0,
        })
    sec_by_id = {s["id"]: s for s in sections}

    # ── Lessons ──────────────────────────────────────────────────────────────
    lessons = []
    les_by_id = {}
    for r in rows(wb["Lessons"]):
        d = {"id": r["lesson_id"], "section_id": r["section_id"], "name": r["lesson_name"],
             "order": r["order"], "total_in_section": 0, "question_count": 0, "question_ids": []}
        lessons.append(d)
        les_by_id[d["id"]] = d

    # ── Question map ─────────────────────────────────────────────────────────
    qrows = sorted(rows(wb["Questions"]), key=lambda r: (r["lesson_id"], r["order_in_lesson"]))
    seen = set()
    for r in qrows:
        qid, lid = r["question_id"], r["lesson_id"]
        if qid in unusable_ids:
            warnings.append(f"question_id {qid} is ungradeable/source-unclear — excluded from practice"); continue
        if qid not in valid_ids:
            warnings.append(f"question_id {qid} not in questions.json — skipped"); continue
        if lid not in les_by_id:
            warnings.append(f"lesson_id {lid} for {qid} not in Lessons sheet — skipped"); continue
        if qid in seen:
            warnings.append(f"question_id {qid} mapped more than once — kept first"); continue
        seen.add(qid)
        les_by_id[lid]["question_ids"].append(qid)

    missing = valid_ids - seen
    if missing:
        warnings.append(f"{len(missing)} question(s) not mapped to any lesson: {sorted(missing)[:8]}…")

    # ── Derive counts ────────────────────────────────────────────────────────
    for l in lessons:
        l["question_count"] = len(l["question_ids"])
    for s in sections:
        sl = [l for l in lessons if l["section_id"] == s["id"]]
        for l in sl:
            l["total_in_section"] = len(sl)
        s["lesson_count"] = len(sl)
        s["question_count"] = sum(l["question_count"] for l in sl)

    out = {"sections": sections, "lessons": lessons}
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    # ── Report ───────────────────────────────────────────────────────────────
    report = ["TOPIC PRACTICE BUILD", "=" * 60,
              f"sections: {len(sections)}  lessons: {len(lessons)}  questions mapped: {len(seen)}", ""]
    if warnings:
        report.append("WARNINGS:")
        report += [f"  ! {w}" for w in warnings]
        report.append("")
    for s in sections:
        report.append(f"{s['name_en']}  ({s['question_count']}q, {s['lesson_count']} lessons, pass {s['pass_correct']}/{s['pass_total']})")
        for l in [x for x in lessons if x["section_id"] == s["id"]]:
            report.append(f"    {l['question_count']:>2}q  {l['name']}")
        report.append("")

    os.makedirs(os.path.dirname(OUT_REPORT), exist_ok=True)
    with open(OUT_REPORT, "w", encoding="utf-8") as f:
        f.write("\n".join(report) + "\n")
    print("\n".join(report))
    print(f"wrote {OUT_JSON}")


if __name__ == "__main__":
    main()
