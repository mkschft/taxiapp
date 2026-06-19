#!/usr/bin/env python3
"""
build_model_tests.py — compile the 5 Model Tests.

Input (SOURCE OF TRUTH):
  content/sources/model_test_workbook.xlsx   clean, ID-keyed
    'Tests'           test_id | title_fi | title_en | time_minutes | pass_mark | order
    'Test_Questions'  test_id | slot | question_id   (bank Qid or MTQ-### id)
    'New_Questions'   id | correct_option | question_fi | question_en |
                      option_*_fi | option_*_en | explanation_en
  src/data/json/questions.json   the shipped bilingual bank (referenced by id)

Provenance: this workbook was frozen from a validated join of the content
editor's raw model-test sheets to the bank, plus translations of the ~80 new
questions that aren't in the bank (see _migrate_model_test_workbook.py). It is
now keyed by stable question_id, so this build is deterministic — to edit a
test, change a cell. The editor's mangled English is never used.

Output:
  src/data/json/model_tests.json           5 tests (id, title, time, pass, question_ids[])
  src/data/json/model_test_questions.json  the NEW questions, Question-shaped,
                                           so getQuestionById resolves them
  content/output/_model_test_report.txt     build summary + integrity warnings

Run from the project root:  python3 content/build_model_tests.py
"""

import json
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
WORKBOOK = os.path.join(HERE, "sources", "model_test_workbook.xlsx")
QUESTIONS = os.path.join(ROOT, "src", "data", "json", "questions.json")
OUT_TESTS = os.path.join(ROOT, "src", "data", "json", "model_tests.json")
OUT_QS = os.path.join(ROOT, "src", "data", "json", "model_test_questions.json")
OUT_REPORT = os.path.join(HERE, "output", "_model_test_report.txt")

KEYS = ("A", "B", "C")

# Official English label per category_id (matches the shipped bank).
CAT_EN = {
    "passenger_safety": "Passenger Help + Safety",
    "special_needs": "Special Passenger Needs",
    "customer_service": "Customer Service",
    "traffic_safety": "Transport + Traffic Safety",
}

# Official Traficom exam split: 15/15/10/10 with per-area minimums. Mocks mirror
# this 1:1 so the per-category pass gate applies. See EXAM_ACCURACY_AUDIT.md.
EXAM_SPLIT = {
    "passenger_safety": 15,
    "special_needs": 15,
    "customer_service": 10,
    "traffic_safety": 10,
}


def rows(ws):
    head = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        yield dict(zip(head, r))


def main():
    bank = json.load(open(QUESTIONS, encoding="utf-8"))
    bank_ids = {q["id"] for q in bank}
    cat_of = {q["id"]: q["category_id"] for q in bank}
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    warnings = []

    # ── New questions (Question-shaped) ──────────────────────────────────────
    new_questions = []
    new_ids = set()
    for r in rows(wb["New_Questions"]):
        new_ids.add(r["id"])
        cid = r.get("category_id")
        if cid not in CAT_EN:
            warnings.append(f"New question {r['id']}: missing/invalid category_id ({cid!r})")
        cat_of[r["id"]] = cid
        new_questions.append({
            "id": r["id"], "category_id": cid, "category_en": CAT_EN.get(cid),
            "source_topic_fi": None, "ref_no": None, "source_set": "model-test",
            "question": {"fi": r["question_fi"], "en": r["question_en"], "fi_raw": None},
            "options": [
                {"key": k, "fi": r[f"option_{k.lower()}_fi"], "fi_raw": None,
                 "en": r[f"option_{k.lower()}_en"], "is_correct": k == r["correct_option"]}
                for k in KEYS
            ],
            "correct_option": r["correct_option"], "correct_master": r["correct_option"],
            "key_overridden": False, "clue_annotations": [],
            "explanation_en": r["explanation_en"], "difficulty": None,
            "tags": ["model-test"], "status": "model-test",
            "fi_edited": False, "reviewer_notes": None, "enriched": False,
        })
    resolvable = bank_ids | new_ids

    # ── Tests + their question lists ─────────────────────────────────────────
    meta = {r["test_id"]: r for r in rows(wb["Tests"])}
    by_test = {tid: [] for tid in meta}
    for r in sorted(rows(wb["Test_Questions"]), key=lambda r: (r["test_id"], r["slot"])):
        qid = r["question_id"]
        if qid not in resolvable:
            warnings.append(f"{r['test_id']} slot {r['slot']}: question_id {qid} not resolvable — skipped")
            continue
        by_test[r["test_id"]].append(qid)

    tests = []
    for tid, m in sorted(meta.items(), key=lambda kv: kv[1]["order"]):
        tests.append({
            "id": tid, "title_fi": m["title_fi"], "title_en": m["title_en"],
            "question_ids": by_test[tid],
            "time_minutes": m["time_minutes"], "pass_mark": m["pass_mark"],
        })

    # ── Integrity: every test = 50 unique questions, split 15/15/10/10 ───────
    # The real Traficom exam is 50 questions across the 4 areas with per-area
    # minimums; mocks must mirror that 1:1 for the pass gate to apply.
    from collections import Counter
    for t in tests:
        ids = t["question_ids"]
        if len(ids) != 50 or len(set(ids)) != 50:
            dups = sorted(q for q in set(ids) if ids.count(q) > 1)
            raise SystemExit(
                f"INTEGRITY ERROR: {t['id']} has {len(ids)} questions "
                f"({len(set(ids))} unique) — expected exactly 50 unique. "
                f"duplicates={dups or 'none'}; fix model_test_workbook.xlsx."
            )
        dist = Counter(cat_of.get(i) for i in ids)
        actual = {c: dist.get(c, 0) for c in EXAM_SPLIT}
        if actual != EXAM_SPLIT:
            raise SystemExit(
                f"INTEGRITY ERROR: {t['id']} category split {actual} "
                f"!= official {EXAM_SPLIT}; fix model_test_workbook.xlsx."
            )

    # ── Integrity: every question used in a test must be gradeable ───────────
    # A question is unusable in a graded test if it has no single correct option
    # or is a flagged "source-unclear" placeholder. Without this guard such a
    # question silently caps the user's score and skews the per-category gate
    # (this is exactly how Q164/Q166 slipped into mt2/mt4 — see the audit).
    rec_of = {q["id"]: q for q in bank}
    rec_of.update({q["id"]: q for q in new_questions})
    ungradeable = []
    for t in tests:
        for qid in t["question_ids"]:
            q = rec_of.get(qid)
            if q is None:
                continue  # unresolvable ids are already reported above
            n_correct = sum(1 for o in q.get("options", []) if o.get("is_correct"))
            if n_correct != 1 or q.get("status") == "source-unclear":
                ungradeable.append(
                    f"{t['id']}: {qid} (status={q.get('status')!r}, "
                    f"correct_options={n_correct})"
                )
    if ungradeable:
        raise SystemExit(
            "INTEGRITY ERROR: ungradeable question(s) used in a model test — "
            "no single correct answer or 'source-unclear' placeholder:\n  "
            + "\n  ".join(ungradeable)
            + "\nFix model_test_workbook.xlsx: swap each for a valid "
            "same-category question."
        )

    with open(OUT_TESTS, "w", encoding="utf-8") as f:
        json.dump(tests, f, ensure_ascii=False, indent=2)
    with open(OUT_QS, "w", encoding="utf-8") as f:
        json.dump(new_questions, f, ensure_ascii=False, indent=2)

    report = ["MODEL TESTS BUILD", "=" * 60,
              f"tests: {len(tests)}  new questions: {len(new_questions)}", ""]
    if warnings:
        report += ["WARNINGS:"] + [f"  ! {w}" for w in warnings] + [""]
    for t in tests:
        nnew = sum(1 for q in t["question_ids"] if q.startswith("MTQ-"))
        report.append(f"  {t['id']}: {len(t['question_ids'])} qs  "
                      f"({len(t['question_ids'])-nnew} bank / {nnew} new)  "
                      f"{t['time_minutes']}min · pass {t['pass_mark']}%")

    os.makedirs(os.path.dirname(OUT_REPORT), exist_ok=True)
    with open(OUT_REPORT, "w", encoding="utf-8") as f:
        f.write("\n".join(report) + "\n")
    print("\n".join(report))
    print(f"\nwrote {OUT_TESTS}\nwrote {OUT_QS}")


if __name__ == "__main__":
    main()
